from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
import sqlite3
import os
from datetime import datetime, timedelta
from functools import wraps
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io

# =========================================
# CONFIGURAÇÃO PRINCIPAL
# =========================================
app = Flask(__name__)

app.config.update(
    SECRET_KEY=os.environ.get("SECRET_KEY", "unirios-sistema-chamados"),
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=30),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=False,       # mudar para True no Render
    SESSION_REFRESH_EACH_REQUEST=True
)
INACTIVITY_TIMEOUT = timedelta(minutes=15)

# =========================================
# CONEXÃO COM BANCO
# =========================================
def conectar():
    os.makedirs(app.instance_path, exist_ok=True)
    db_path = os.path.join(app.instance_path, "chamados.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def inicializar_banco():
    conn = conectar()
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS chamados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        local TEXT NOT NULL,
        setor_responsavel TEXT NOT NULL,
        descricao TEXT NOT NULL,
        prioridade TEXT NOT NULL,
        status TEXT DEFAULT 'aberto',
        criado_em TEXT NOT NULL
    )
    """)
    conn.commit()
    conn.close()

inicializar_banco()

# =========================================
# LOGIN DOS SETORES
# =========================================
setores = {
    "NTI": {"usuario": "nti", "senha": "nti123"},
    "Elétrica/Clima": {"usuario": "eletrica", "senha": "eletrica123"},
    "Almoxarifado": {"usuario": "almox", "senha": "almox123"},
    "Limpeza": {"usuario": "limpeza", "senha": "limpeza123"},
    "Manutenção Predial": {"usuario": "predial", "senha": "predial123"}
}

# =========================================
# FUNÇÕES DE SEGURANÇA
# =========================================
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped

@app.before_request
def enforce_session_timeout():
    if "usuario" not in session:
        return
    now = datetime.utcnow().timestamp()
    last = session.get("last_seen", now)
    if now - last > INACTIVITY_TIMEOUT.total_seconds():
        session.clear()
        return redirect(url_for("login"))
    session["last_seen"] = now

@app.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0, private"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

def is_nti():
    return session.get("usuario") == "nti"

# =========================================
# ROTAS PÚBLICAS
# =========================================
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/novo')
def novo_chamado():
    return render_template('novo.html', setores_responsaveis=list(setores.keys()))

@app.route('/abrir', methods=['POST'])
def abrir():
    local = request.form['local'].strip()
    setor_responsavel = request.form['setor_responsavel'].strip()
    descricao = request.form['descricao'].strip()
    prioridade = request.form['prioridade'].strip()
    criado_em = datetime.now().strftime("%d/%m/%Y %H:%M")

    conn = conectar()
    c = conn.cursor()
    c.execute("""
        INSERT INTO chamados (local, setor_responsavel, descricao, prioridade, criado_em)
        VALUES (?, ?, ?, ?, ?)
    """, (local, setor_responsavel, descricao, prioridade, criado_em))
    conn.commit()
    protocolo = c.lastrowid
    conn.close()

    return redirect(url_for('sucesso', protocolo=protocolo))

@app.route('/sucesso/<int:protocolo>')
def sucesso(protocolo):
    return render_template('sucesso.html', protocolo=protocolo)

# =========================================
# LOGIN / LOGOUT
# =========================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        usuario = request.form['usuario'].strip()
        senha = request.form['senha'].strip()
        for setor, cred in setores.items():
            if cred["usuario"] == usuario and cred["senha"] == senha:
                session.clear()
                session.permanent = True
                session['usuario'] = usuario
                session['setor'] = setor
                session['last_seen'] = datetime.utcnow().timestamp()
                return redirect(url_for('painel'))
        erro = "Setor ou senha inválidos."
    return render_template('login.html', erro=erro)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# =========================================
# PAINEL TÉCNICO (com filtro por setor)
# =========================================
@app.route('/painel')
@login_required
def painel():
    setor = session.get('setor')
    status = request.args.get('status', 'aberto')
    setor_filtro = request.args.get('setor_filtro')  # novo filtro

    conn = conectar()
    c = conn.cursor()

    # NTI pode ver tudo e filtrar por setor
    if is_nti():
        query = "SELECT * FROM chamados"
        params = []

        filtros = []
        if status != 'todos':
            filtros.append("status = ?")
            params.append(status)
        if setor_filtro:
            filtros.append("setor_responsavel = ?")
            params.append(setor_filtro)

        if filtros:
            query += " WHERE " + " AND ".join(filtros)
        query += " ORDER BY id DESC"

        c.execute(query, params)
    else:
        if status == 'todos':
            c.execute("SELECT * FROM chamados WHERE setor_responsavel = ? ORDER BY id DESC", (setor,))
        else:
            c.execute("SELECT * FROM chamados WHERE setor_responsavel = ? AND status = ? ORDER BY id DESC", (setor, status))

    chamados = c.fetchall()
    conn.close()

    return render_template('painel.html', chamados=chamados, setor=setor, filtro=status)

# =========================================
# ALTERAR STATUS / EXCLUIR
# =========================================
@app.route('/alterar/<int:cid>', methods=['POST'])
@login_required
def alterar(cid):
    novo_status = request.form['status']
    conn = conectar()
    c = conn.cursor()
    row = c.execute("SELECT setor_responsavel FROM chamados WHERE id = ?", (cid,)).fetchone()
    if not row:
        conn.close()
        return redirect(url_for('painel'))
    setor_resp = row['setor_responsavel']
    if not is_nti() and setor_resp != session.get('setor'):
        conn.close()
        return redirect(url_for('painel'))
    c.execute("UPDATE chamados SET status = ? WHERE id = ?", (novo_status, cid))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('painel'))

@app.route('/excluir/<int:cid>', methods=['POST'])
@login_required
def excluir(cid):
    if session.get('usuario') != 'nti':
        flash('Apenas o NTI pode excluir chamados.', 'danger')
        return redirect(url_for('painel'))
    conn = conectar()
    c = conn.cursor()
    c.execute("DELETE FROM chamados WHERE id = ?", (cid,))
    conn.commit()
    conn.close()
    flash('Chamado excluído com sucesso!', 'success')
    return redirect(request.referrer or url_for('painel'))

# =========================================
# DASHBOARD NTI
# =========================================
@app.route('/dashboard')
@login_required
def dashboard():
    if not is_nti():
        return redirect(url_for('painel'))
    conn = conectar()
    c = conn.cursor()
    c.execute("SELECT * FROM chamados ORDER BY id DESC LIMIT 50")
    ultimos = c.fetchall()
    conn.close()
    return render_template('dashboard.html', ultimos=ultimos)

def parse_data_ptbr(dt_str):
    try:
        return datetime.strptime(dt_str, "%d/%m/%Y %H:%M")
    except Exception:
        return None

@app.route('/api/dashboard/overview')
@login_required
def api_overview():
    if not is_nti(): return jsonify({}), 403
    conn = conectar()
    c = conn.cursor()
    totals = {
        "total": c.execute("SELECT COUNT(*) FROM chamados").fetchone()[0],
        "abertos": c.execute("SELECT COUNT(*) FROM chamados WHERE status='aberto'").fetchone()[0],
        "andamento": c.execute("SELECT COUNT(*) FROM chamados WHERE status='em andamento'").fetchone()[0],
        "finalizados": c.execute("SELECT COUNT(*) FROM chamados WHERE status='finalizado'").fetchone()[0]
    }
    conn.close()
    return jsonify(totals)

@app.route('/api/dashboard/by_sector')
@login_required
def api_by_sector():
    if not is_nti(): return jsonify({}), 403
    conn = conectar()
    c = conn.cursor()
    rows = c.execute("""
        SELECT setor_responsavel, COUNT(*) as total
        FROM chamados
        GROUP BY setor_responsavel
        ORDER BY total DESC
    """).fetchall()
    conn.close()
    labels = [r["setor_responsavel"] for r in rows]
    values = [r["total"] for r in rows]
    return jsonify({"labels": labels, "values": values})

@app.route('/api/dashboard/by_day')
@login_required
def api_by_day():
    if not is_nti(): return jsonify({}), 403
    limite = datetime.now() - timedelta(days=29)
    buckets = defaultdict(int)
    conn = conectar()
    c = conn.cursor()
    rows = c.execute("SELECT criado_em FROM chamados").fetchall()
    conn.close()
    for r in rows:
        dt = parse_data_ptbr(r["criado_em"])
        if dt and dt.date() >= limite.date():
            buckets[dt.strftime("%d/%m")] += 1
    dias = [(limite + timedelta(days=i)).strftime("%d/%m") for i in range(30)]
    values = [buckets.get(d, 0) for d in dias]
    return jsonify({"labels": dias, "values": values})

# =========================================
# BACKUP EXCEL (ÚLTIMOS 30 DIAS)
# =========================================
@app.route('/backup')
@login_required
def backup_excel():
    if not is_nti():
        flash("Apenas o NTI pode realizar backups.", "danger")
        return redirect(url_for("painel"))

    # Limite de 30 dias atrás
    limite_data = datetime.now() - timedelta(days=30)

    conn = conectar()
    c = conn.cursor()
    c.execute("SELECT * FROM chamados ORDER BY id DESC")
    todos_chamados = c.fetchall()
    conn.close()

    # Filtrar apenas chamados dos últimos 30 dias
    chamados_filtrados = []
    for c_row in todos_chamados:
        try:
            data_criacao = datetime.strptime(c_row["criado_em"], "%d/%m/%Y %H:%M")
            if data_criacao >= limite_data:
                chamados_filtrados.append(c_row)
        except Exception:
            continue

    # Agrupar corretamente por setor
    agrupados_por_setor = {}
    for chamado in chamados_filtrados:
        setor = chamado["setor_responsavel"]
        if setor not in agrupados_por_setor:
            agrupados_por_setor[setor] = []
        agrupados_por_setor[setor].append(chamado)

    # Criar planilha Excel
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    invalid_chars = set(r'[]:*?/\\')

    for setor, lista in agrupados_por_setor.items():
        nome_sheet = ''.join(ch for ch in setor if ch not in invalid_chars and ch.isprintable()).strip()[:30]
        ws = wb.create_sheet(title=nome_sheet or "Setor")

        headers = ["ID", "Setor Responsável", "Local", "Descrição", "Prioridade", "Status", "Criado em"]
        ws.append(headers)

        # Cabeçalho estilizado
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Dados
        for c_row in lista:
            ws.append([
                c_row["id"],
                c_row["setor_responsavel"],
                c_row["local"],
                c_row["descricao"],
                c_row["prioridade"],
                c_row["status"].capitalize(),
                c_row["criado_em"]
            ])

        # Ajustar largura das colunas
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Caso não haja chamados no período
    if not agrupados_por_setor:
        ws = wb.create_sheet(title="Sem chamados")
        ws.append(["Nenhum chamado foi registrado nos últimos 30 dias."])
        ws.column_dimensions["A"].width = 60

    # Salvar arquivo em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"backup_chamados_ultimos30dias_{datetime.now().strftime('%Y-%m-%d_%Hh%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================================
# EXECUÇÃO LOCAL
# =========================================
if __name__ == '__main__':
    app.run(debug=True)
